"""
dashboard.py
Generates a beautiful HTML dashboard + enhanced Excel from IHC analysis results.
Called from run_pipeline.py after all images are processed.
"""

import os
from pathlib import Path
from datetime import datetime


# ==========================================================
# HTML DASHBOARD
# ==========================================================

def generate_html_dashboard(batch_metrics: list, output_path: str, config: dict = None):
    """
    Generate a beautiful dark-themed HTML dashboard from batch results.
    Opens in any browser, no dependencies needed.
    """

    if not batch_metrics:
        return

    total_cells = sum(m["Total_Cells"] for m in batch_metrics)
    total_positive = sum(m["Positive_Cells"] for m in batch_metrics)
    avg_positivity = sum(m["Positivity_Index_Pct"] for m in batch_metrics) / len(batch_metrics)
    avg_cells = total_cells // len(batch_metrics)

    labels_js = str([m["Image_Name"].split(" - ")[0] for m in batch_metrics])
    neg_js = str([m["Negative_Cells"] for m in batch_metrics])
    pos_js = str([m["Positive_Cells"] for m in batch_metrics])
    pct_js = str([m["Positivity_Index_Pct"] for m in batch_metrics])
    total_js = str([m["Total_Cells"] for m in batch_metrics])

    max_pct = max(m["Positivity_Index_Pct"] for m in batch_metrics)

    table_rows = ""
    for m in batch_metrics:
        name = m["Image_Name"].split(" - ")[0]
        bar_width = min(int((m["Positivity_Index_Pct"] / 10.0) * 100), 100)
        confidence_class = "badge-normal" if m["Confidence"] == "NORMAL" else "badge-low"
        table_rows += f"""
        <tr>
            <td>{name}</td>
            <td>{m['Total_Cells']:,}</td>
            <td class="pos-val">{m['Positive_Cells']:,}</td>
            <td>{m['Negative_Cells']:,}</td>
            <td>{m['Positivity_Index_Pct']:.2f}%</td>
            <td style="min-width:100px">
                <div class="pos-bar"><div class="pos-fill" style="width:{bar_width}%"></div></div>
<div style="font-size:10px;color:#475569;margin-top:3px">0 — 10% scale</div>
            </td>
            <td>{m.get('Pixel_Size_um', 0.5)} µm/px</td>
            <td><span class="badge {confidence_class}">{m['Confidence']}</span></td>
        </tr>"""

    threshold = config.get("dab_threshold", 0.2) if config else 0.2
    stain = config.get("stain_type", "H-DAB").upper() if config else "H-DAB"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IHC Analysis Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #0f1117;
    color: #e2e8f0;
    min-height: 100vh;
    padding: 2rem;
  }}

  .container {{ max-width: 1200px; margin: 0 auto; }}

  .header {{ margin-bottom: 2rem; }}
  .header h1 {{ font-size: 22px; font-weight: 500; color: #f1f5f9; margin-bottom: 4px; }}
  .header p {{ font-size: 13px; color: #64748b; }}

  .metrics {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin-bottom: 1.5rem;
  }}

  .metric {{
    background: #1e2130;
    border: 0.5px solid #2d3348;
    border-radius: 10px;
    padding: 16px 18px;
  }}

  .metric-label {{
    font-size: 11px;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-bottom: 8px;
  }}

  .metric-value {{
    font-size: 26px;
    font-weight: 500;
    color: #f1f5f9;
    line-height: 1;
  }}

  .metric-value.accent {{ color: #f97316; }}
  .metric-sub {{ font-size: 12px; color: #475569; margin-top: 4px; }}

  .charts-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
    margin-bottom: 16px;
  }}

  .chart-card {{
    background: #1e2130;
    border: 0.5px solid #2d3348;
    border-radius: 10px;
    padding: 18px;
  }}

  .chart-title {{
    font-size: 11px;
    font-weight: 500;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-bottom: 14px;
  }}

  .legend {{
    display: flex;
    gap: 16px;
    margin-bottom: 10px;
    font-size: 11px;
    color: #64748b;
    flex-wrap: wrap;
  }}

  .legend span {{
    display: flex;
    align-items: center;
    gap: 5px;
  }}

  .legend-dot {{
    width: 10px;
    height: 10px;
    border-radius: 2px;
    flex-shrink: 0;
  }}

  .table-card {{
    background: #1e2130;
    border: 0.5px solid #2d3348;
    border-radius: 10px;
    padding: 18px;
    margin-bottom: 16px;
    overflow-x: auto;
  }}

  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}

  th {{
    text-align: left;
    padding: 8px 12px;
    color: #64748b;
    font-weight: 500;
    border-bottom: 0.5px solid #2d3348;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    white-space: nowrap;
  }}

  td {{
    padding: 12px 12px;
    border-bottom: 0.5px solid #1a1f2e;
    color: #cbd5e1;
    white-space: nowrap;
  }}

  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: rgba(255,255,255,0.02); }}

  .pos-val {{ color: #f97316; font-weight: 500; }}

  .pos-bar {{
    height: 6px;
    background: #2d3348;
    border-radius: 3px;
    overflow: hidden;
  }}

  .pos-fill {{
    height: 100%;
    border-radius: 3px;
    background: #f97316;
    transition: width 0.3s;
  }}

  .badge {{
    display: inline-block;
    padding: 3px 8px;
    border-radius: 4px;
    font-size: 11px;
    font-weight: 500;
  }}

  .badge-normal {{ background: #14532d; color: #86efac; }}
  .badge-low {{ background: #7f1d1d; color: #fca5a5; }}

  .footer {{
    margin-top: 2rem;
    font-size: 12px;
    color: #334155;
    text-align: center;
  }}

  @media (max-width: 768px) {{
    .metrics {{ grid-template-columns: repeat(2, 1fr); }}
    .charts-grid {{ grid-template-columns: 1fr; }}
  }}
</style>
</head>
<body>
<div class="container">

  <div class="header">
    <h1>IHC Analysis Dashboard</h1>
    <p>{len(batch_metrics)} images · {stain} · threshold {threshold} OD · {timestamp}</p>
  </div>

  <div class="metrics">
    <div class="metric">
      <div class="metric-label">Total cells detected</div>
      <div class="metric-value">{total_cells:,}</div>
      <div class="metric-sub">{len(batch_metrics)} images analyzed</div>
    </div>
    <div class="metric">
      <div class="metric-label">DAB positive</div>
      <div class="metric-value accent">{total_positive:,}</div>
      <div class="metric-sub">{avg_positivity:.2f}% avg positivity</div>
    </div>
    <div class="metric">
      <div class="metric-label">Avg cells / image</div>
      <div class="metric-value">{avg_cells:,}</div>
      <div class="metric-sub">range varies by tissue</div>
    </div>
    <div class="metric">
      <div class="metric-label">Confidence</div>
      <div class="metric-value">{"✓" if all(m["Confidence"] == "NORMAL" for m in batch_metrics) else "⚠"}</div>
      <div class="metric-sub">{"All normal" if all(m["Confidence"] == "NORMAL" for m in batch_metrics) else "Review flagged images"}</div>
    </div>
  </div>

  <div class="charts-grid">
    <div class="chart-card">
      <div class="chart-title">Cell counts per image</div>
      <div class="legend">
        <span><span class="legend-dot" style="background:#3b82f6"></span>Negative</span>
        <span><span class="legend-dot" style="background:#f97316"></span>Positive</span>
      </div>
      <div style="position:relative;height:220px">
        <canvas id="c1" role="img" aria-label="Stacked bar chart showing negative and positive cell counts per image">Stacked cell counts across {len(batch_metrics)} images.</canvas>
      </div>
    </div>
    <div class="chart-card">
      <div class="chart-title">DAB positivity rate</div>
      <div class="legend">
        <span><span class="legend-dot" style="background:#f97316"></span>Positivity %</span>
      </div>
      <div style="position:relative;height:220px">
        <canvas id="c2" role="img" aria-label="Bar chart of DAB positivity percentage per image">Positivity rates across images.</canvas>
      </div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Cell density per image</div>
      <div style="position:relative;height:220px">
        <canvas id="c3" role="img" aria-label="Horizontal bar chart of total cell counts per image">Cell density comparison across images.</canvas>
      </div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Positive cells vs total cells</div>
      <div style="position:relative;height:220px">
        <canvas id="c4" role="img" aria-label="Scatter plot of positive cell count versus total cells">Correlation between total cells and positive cells.</canvas>
      </div>
    </div>
  </div>

  <div class="table-card">
    <div class="chart-title" style="margin-bottom:14px">Per image results</div>
    <table>
      <thead>
        <tr>
          <th>Image</th>
          <th>Total cells</th>
          <th>Positive</th>
          <th>Negative</th>
          <th>Positivity %</th>
          <th>Positivity bar</th>
          <th>Pixel size</th>
          <th>Confidence</th>
        </tr>
      </thead>
      <tbody>{table_rows}</tbody>
    </table>
  </div>

  <div class="footer">
    Generated by IHC Analyzer · InstanSeg + QuPath · DAB threshold {threshold} OD
  </div>

</div>
<script>
const labels = {labels_js};
const neg = {neg_js};
const pos = {pos_js};
const pct = {pct_js};
const total = {total_js};

const grid = 'rgba(255,255,255,0.06)';
const tick = '#475569';

const base = {{
  responsive: true,
  maintainAspectRatio: false,
  plugins: {{ legend: {{ display: false }}, tooltip: {{ cornerRadius: 6, backgroundColor: '#1e2130', borderColor: '#2d3348', borderWidth: 1 }} }},
  scales: {{
    x: {{ grid: {{ color: grid }}, ticks: {{ color: tick, font: {{ size: 11 }} }} }},
    y: {{ grid: {{ color: grid }}, ticks: {{ color: tick, font: {{ size: 11 }} }} }}
  }}
}};

new Chart(document.getElementById('c1'), {{
  type: 'bar',
  data: {{
    labels,
    datasets: [
      {{ label: 'Negative', data: neg, backgroundColor: '#3b82f6', borderRadius: 3 }},
      {{ label: 'Positive', data: pos, backgroundColor: '#f97316', borderRadius: 3 }}
    ]
  }},
  options: {{ ...base, scales: {{ x: {{ ...base.scales.x, stacked: true }}, y: {{ ...base.scales.y, stacked: true }} }} }}
}});

new Chart(document.getElementById('c2'), {{
  type: 'bar',
  data: {{
    labels,
    datasets: [{{ label: 'Positivity %', data: pct, backgroundColor: '#f97316', borderRadius: 3 }}]
  }},
  options: {{ ...base, scales: {{ x: base.scales.x, y: {{ ...base.scales.y, ticks: {{ ...base.scales.y.ticks, callback: v => v + '%' }} }} }} }}
}});

new Chart(document.getElementById('c3'), {{
  type: 'bar',
  data: {{
    labels: [...labels].reverse(),
    datasets: [{{ label: 'Total cells', data: [...total].reverse(), backgroundColor: '#3b82f6', borderRadius: 3 }}]
  }},
  options: {{ ...base, indexAxis: 'y', scales: {{ x: {{ ...base.scales.x, ticks: {{ ...base.scales.x.ticks, callback: v => v >= 1000 ? (v/1000).toFixed(0)+'k' : v }} }}, y: base.scales.y }} }}
}});

new Chart(document.getElementById('c4'), {{
  type: 'scatter',
  data: {{
    datasets: [{{
      label: 'Images',
      data: total.map((t,i) => ({{ x: t, y: pos[i] }})),
      backgroundColor: '#f97316',
      pointRadius: 8,
      pointHoverRadius: 10
    }}]
  }},
  options: {{ ...base, scales: {{
    x: {{ ...base.scales.x, title: {{ display: true, text: 'Total cells', color: tick, font: {{ size: 11 }} }} }},
    y: {{ ...base.scales.y, title: {{ display: true, text: 'Positive cells', color: tick, font: {{ size: 11 }} }} }}
  }} }}
}});
</script>
</body>
</html>"""

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"  Dashboard: {output_path}")
    return output_path


# ==========================================================
# EXCEL EXPORT
# ==========================================================

def generate_excel(batch_metrics: list, output_path: str, config: dict = None):
    """
    Generate a multi-sheet Excel with:
    - Sheet 1: Batch Summary
    - Sheet 2: Per Image Details
    - Sheet 3: Statistics
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("  Installing openpyxl...")
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Colors ────────────────────────────────────────────
    DARK_BLUE  = "1E3A5F"
    MED_BLUE   = "2D6A9F"
    LIGHT_BLUE = "D6E4F0"
    ORANGE     = "E8650A"
    LIGHT_ORANGE = "FDE9D9"
    GRAY_BG    = "F5F5F5"
    WHITE      = "FFFFFF"
    DARK_TEXT  = "1A1A1A"

    def header_style(cell, bg=DARK_BLUE, fg=WHITE):
        cell.font = Font(bold=True, color=fg, size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def data_style(cell, bg=WHITE):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=10, color=DARK_TEXT)

    def thin_border():
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Batch Summary ────────────────────────────
    ws1 = wb.active
    ws1.title = "Batch Summary"
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = "IHC Analysis — Batch Summary"
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 32

    # Subtitle
    ws1.merge_cells("A2:H2")
    sub = ws1["A2"]
    threshold = config.get("dab_threshold", 0.2) if config else 0.2
    stain = config.get("stain_type", "H-DAB").upper() if config else "H-DAB"
    sub.value = f"{stain} · DAB threshold {threshold} OD · {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub.font = Font(size=10, color="666666")
    sub.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 20

    ws1.row_dimensions[3].height = 8  # spacer

    # Headers
    headers = ["Image", "Total Cells", "Positive", "Negative",
               "Positivity %", "DAB Threshold", "Pixel Size (µm)", "Confidence"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        header_style(cell, MED_BLUE)
        cell.border = thin_border()
    ws1.row_dimensions[4].height = 22

    # Data
    for row_i, m in enumerate(batch_metrics, 5):
        bg = WHITE if row_i % 2 == 1 else GRAY_BG
        values = [
            m["Image_Name"].split(" - ")[0],
            m["Total_Cells"],
            m["Positive_Cells"],
            m["Negative_Cells"],
            m["Positivity_Index_Pct"] / 100,
            m.get("DAB_Threshold", 0.2),
            m.get("Pixel_Size_um", 0.5),
            m["Confidence"]
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_i, column=col, value=val)
            data_style(cell, bg)
            cell.border = thin_border()
            if col == 3:  # Positive
                cell.font = Font(size=10, color=ORANGE, bold=True)
            if col == 5:  # Positivity %
                cell.number_format = "0.00%"
            if col == 8:  # Confidence
                if val == "NORMAL":
                    cell.fill = PatternFill("solid", fgColor="D4EDDA")
                    cell.font = Font(size=10, color="155724")
                else:
                    cell.fill = PatternFill("solid", fgColor="F8D7DA")
                    cell.font = Font(size=10, color="721C24")
        ws1.row_dimensions[row_i].height = 20

    # Column widths
    col_widths = [35, 14, 12, 12, 14, 15, 17, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Statistics ───────────────────────────────
    ws2 = wb.create_sheet("Statistics")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:C1")
    ws2["A1"].value = "Batch Statistics"
    ws2["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws2["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 32

    total_cells = sum(m["Total_Cells"] for m in batch_metrics)
    total_pos = sum(m["Positive_Cells"] for m in batch_metrics)
    total_neg = sum(m["Negative_Cells"] for m in batch_metrics)
    avg_pos_pct = sum(m["Positivity_Index_Pct"] for m in batch_metrics) / len(batch_metrics)
    min_pct = min(m["Positivity_Index_Pct"] for m in batch_metrics)
    max_pct = max(m["Positivity_Index_Pct"] for m in batch_metrics)

    stats = [
        ("Images analyzed", len(batch_metrics)),
        ("Total cells detected", total_cells),
        ("Total DAB positive", total_pos),
        ("Total DAB negative", total_neg),
        ("Average positivity %", f"{avg_pos_pct:.2f}%"),
        ("Min positivity %", f"{min_pct:.2f}%"),
        ("Max positivity %", f"{max_pct:.2f}%"),
        ("DAB threshold (OD)", threshold),
        ("Analysis date", datetime.now().strftime("%Y-%m-%d")),
    ]

    for row_i, (label, value) in enumerate(stats, 3):
        bg = WHITE if row_i % 2 == 1 else GRAY_BG
        lc = ws2.cell(row=row_i, column=1, value=label)
        lc.font = Font(size=10, bold=True, color=DARK_TEXT)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.border = thin_border()
        lc.alignment = Alignment(horizontal="left", indent=1)

        vc = ws2.cell(row=row_i, column=2, value=value)
        vc.font = Font(size=10, color=DARK_TEXT)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.border = thin_border()
        vc.alignment = Alignment(horizontal="center")
        ws2.row_dimensions[row_i].height = 20

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    # ── Sheet 3: Raw Data ─────────────────────────────────
    ws3 = wb.create_sheet("Raw Data")
    ws3.sheet_view.showGridLines = False

    raw_headers = ["Image", "Total_Cells", "Positive_Cells", "Negative_Cells",
                   "Positivity_Pct", "DAB_Threshold", "Pixel_Size_um", "Confidence"]
    for col, h in enumerate(raw_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        header_style(cell, DARK_BLUE)
        cell.border = thin_border()

    for row_i, m in enumerate(batch_metrics, 2):
        row_data = [
            m["Image_Name"],
            m["Total_Cells"],
            m["Positive_Cells"],
            m["Negative_Cells"],
            m["Positivity_Index_Pct"],
            m.get("DAB_Threshold", 0.2),
            m.get("Pixel_Size_um", 0.5),
            m["Confidence"]
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_i, column=col, value=val)
            cell.font = Font(size=10)
            cell.border = thin_border()

    for i, w in enumerate([50, 14, 14, 14, 14, 14, 14, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"  Excel:     {output_path}")
    return output_path


# ==========================================================
# MAIN ENTRY — called from run_pipeline.py
# ==========================================================

def generate_all_outputs(batch_metrics: list, dashboard_dir: str, config: dict = None):
    """Generate HTML dashboard + Excel from batch results."""

    os.makedirs(dashboard_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    html_path = os.path.join(dashboard_dir, f"ihc_dashboard_{timestamp}.html")
    excel_path = os.path.join(dashboard_dir, f"ihc_results_{timestamp}.xlsx")

    generate_html_dashboard(batch_metrics, html_path, config)
    generate_excel(batch_metrics, excel_path, config)

    return html_path, excel_path